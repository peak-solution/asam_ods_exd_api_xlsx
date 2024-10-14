import logging
import pathlib
import tempfile
import unittest
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook


def calculate_mean_of_string_lengths(meta_rows):
    mean_lengths = {}
    for index, row in meta_rows.iterrows():
        string_values = row.apply(
            lambda x: x if isinstance(x, str) else None).dropna()
        if not string_values.empty:
            mean_length = string_values.apply(len).mean()
            mean_lengths[index] = mean_length
        else:
            mean_lengths[index] = None
    return mean_lengths


def find_stable_datatype_row(df):
    # Get the datatypes of the first row
    previous_dtypes = df.iloc[0].apply(type)

    for index, row in df.iterrows():
        current_dtypes = row.apply(type)

        # Check if current row datatypes match the previous row datatypes
        if not current_dtypes.equals(previous_dtypes):
            previous_dtypes = current_dtypes
        else:
            # Check if at least one column isn't a string
            if any(dtype != str for dtype in current_dtypes):
                return index - 1  # Return the index of the previous row

    return None  # If no consistent row is found


def extract_excel_information(file_path):
    with pd.ExcelFile(file_path) as xls:
        for sheet_index, sheet_name in enumerate(xls.sheet_names):
            df = pd.read_excel(xls, sheet_name=sheet_name,
                               nrows=100, skiprows=0)

            df_by_index = pd.read_excel(xls, sheet_name=sheet_index,
                                        nrows=100, skiprows=0)

            assert df.equals(df_by_index)

            first_stable_row = find_stable_datatype_row(df)

            assert first_stable_row is not None, "No stable row found"
            print(f"First stable row: {first_stable_row}")

            meta_rows = pd.DataFrame(
                columns=df.columns) if first_stable_row == 0 else df.iloc[:first_stable_row - 1].copy()
            print(meta_rows)

            mean_lengths = calculate_mean_of_string_lengths(meta_rows)
            print(f"Mean lengths of strings in meta rows: {mean_lengths}")

            # Drop till stable row
            df.drop(range(first_stable_row - 1), inplace=True)
            df.reset_index(drop=True, inplace=True)
            df = df.infer_objects()

            # Print the DataFrame to verify
            print(df)

            # Check if the first column is monotonically increasing
            is_monotonic = df.iloc[:, 0].is_monotonic_increasing
            assert is_monotonic, "The first column is not monotonically increasing"

            # Read rows 5 to 10 and columns 7 to 10
            df_subset = pd.read_excel(
                file_path, sheet_name="data_1", usecols=[0, 1, 3], skiprows=first_stable_row + 1, nrows=2)

            # Print the subset DataFrame to verify
            print(df_subset)


class TestStringMethods(unittest.TestCase):
    log = logging.getLogger(__name__)

    def _get_example_file_path(self, file_name):
        example_file_path = pathlib.Path.joinpath(pathlib.Path(
            __file__).parent.resolve(), "..", "data", file_name)
        return pathlib.Path(example_file_path).absolute()

    def test_create_file_with_unit(self):
        with tempfile.TemporaryDirectory() as tmp_dir_name:
            file_path = pathlib.Path.joinpath(
                pathlib.Path(tmp_dir_name), "example_data_with_unit.xlsx")
            now = datetime.now()
            wb = Workbook()
            ws = wb.active
            ws.title = "data_1"
            ws.append(["time", "index", "speed", "comment"])
            ws.append(["s", "-", "m/s", "-"])
            ws.append([now + timedelta(seconds=0), 1, 3.0, "abc"])
            ws.append([now + timedelta(seconds=1), 2, 3.1, "def"])
            ws.append([now + timedelta(seconds=2), 3, 3.2, "ghi"])
            wb.save(file_path)
            wb.close()

            extract_excel_information(file_path)

    def test_create_file_with_unit_and_comment(self):
        with tempfile.TemporaryDirectory() as tmp_dir_name:
            file_path = pathlib.Path.joinpath(
                pathlib.Path(tmp_dir_name), "example_data_with_unit_and_comment.xlsx")
            now = datetime.now()
            wb = Workbook()
            ws = wb.active
            ws.title = "data_1"
            ws.append(["time", "index", "speed", "comment"])
            ws.append(["s", "-", "m/s", "-"])
            ws.append(["time channel", "index",
                      "Speed of the vehicle", "comment column"])
            ws.append([now + timedelta(seconds=0), 1, 3.0, "abc"])
            ws.append([now + timedelta(seconds=1), 2, 3.1, "def"])
            ws.append([now + timedelta(seconds=2), 3, 3.2, "ghi"])
            wb.save(file_path)
            wb.close()

            extract_excel_information(file_path)

    def test_create_file(self):
        with tempfile.TemporaryDirectory() as tmp_dir_name:
            file_path = pathlib.Path.joinpath(
                pathlib.Path(tmp_dir_name), "example_data.xlsx")
            now = datetime.now()
            wb = Workbook()
            ws = wb.active
            ws.title = "data_1"
            ws.append(["time", "index", "speed", "comment"])
            ws.append([now + timedelta(seconds=0), 1, 3.0, "abc"])
            ws.append([now + timedelta(seconds=1), 2, 3.1, "def"])
            ws.append([now + timedelta(seconds=2), 3, 3.2, "ghi"])
            # Save the file
            wb.save(file_path)
            wb.close()

            extract_excel_information(file_path)

    def test_create_file_multi_sheet(self):
        with tempfile.TemporaryDirectory() as tmp_dir_name:
            file_path = pathlib.Path.joinpath(
                pathlib.Path(tmp_dir_name), "example_data_two_sheets.xlsx")

            now = datetime.now()
            wb = Workbook()
            ws = wb.active
            ws.title = "data_1"
            ws.append(["time", "index", "speed", "comment"])
            ws.append([now + timedelta(seconds=0), 1, 3.0, "abc"])
            ws.append([now + timedelta(seconds=1), 2, 3.1, "def"])
            ws.append([now + timedelta(seconds=2), 3, 3.2, "ghi"])
            ws2 = wb.create_sheet("data_2")
            ws2.append(["time", "index", "speed", "comment"])
            ws2.append([now + timedelta(seconds=3), 4, 4.0, "abc2"])
            ws2.append([now + timedelta(seconds=4), 5, 4.1, "def2"])
            ws2.append([now + timedelta(seconds=5), 6, 4.2, "ghi2"])
            ws2.append([now + timedelta(seconds=6), 7, 4.3, "jkl2"])
            ws2.append([now + timedelta(seconds=7), 8, 4.4, "mno2"])
            wb.save(file_path)
            wb.close()

            extract_excel_information(file_path)
